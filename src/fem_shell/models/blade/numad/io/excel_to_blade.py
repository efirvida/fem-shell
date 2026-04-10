"""
Read a NuMAD Excel blade definition file (.xlsx) into a Blade object.

Adapted from pyNuMAD (Sandia National Laboratories) ``excel_to_blade.py``.
Uses openpyxl instead of pandas to avoid a heavy dependency.

Two Excel formats are supported:

**New (pyNuMAD) format** – three sheets:

* **Geometry** – span, twist, chord, percent-thick, chord-offset,
  aerocenter, airfoil span-locations, airfoil names, and interpolated
  span stations.  Configuration flags in rows 2-4.
* **Components** – structural layer definitions with group, name,
  material-id, fabric-angle, HP/LP extents, control-point span/n-layers,
  and interpolation method.
* **Materials** – material property table.

**Legacy (station-based) NuMAD format** – five sheets:

* **Geometry** – station-based layout with per-station geometry, stack
  layer counts, segment delineation points, DP types, included-stacks
  per segment, and shear-web stacks/connections.
* **Materials** – same structure as the new format, but data starts at
  row 4 and has no dry-density column.
* **bend & sweep** – prebend and sweep distributions.
* **SW** – spar-cap and shear-web keypoint positions per station.
* **BOM** – bill of materials (mass summary, not used by the reader).

Format detection is automatic: the presence of a ``Components`` sheet
selects the new format.
"""

import logging
import os

import numpy as np
from openpyxl import load_workbook

from fem_shell.models.blade.numad.objects.airfoil import Airfoil
from fem_shell.models.blade.numad.objects.component import Component
from fem_shell.models.blade.numad.objects.definition import Definition
from fem_shell.models.blade.numad.objects.material import Material
from fem_shell.models.blade.numad.utils.interpolation import interpolator_wrap

logger = logging.getLogger(__name__)

# Fixed column indices (0-based) for each sheet – mirrors the NuMAD Excel layout.
_GEOM = {
    "datarow1": 7,  # first data row (1-indexed in Excel → row 7, row 6 is headers)
    "span": 0,
    "twist": 1,
    "chord": 2,
    "thick": 3,
    "offset": 4,
    "aerocenter": 5,
    "afspan": 7,
    "afname": 8,
    "ispan": 10,
}

_CMPT = {
    "paramcol": 2,
    "paramrow1": 1,  # 1-indexed
    "datarow1": 7,
    "group": 0,
    "name": 1,
    "matid": 2,
    "angle": 3,
    "hpext": 4,
    "lpext": 5,
    "cpspan": 6,
    "cpnlay": 7,
    "imethod": 8,
}

_MTRL = {
    "datarow1": 4,
    "id": 0,
    "name": 1,
    "type": 2,
    "thickness": 3,
    "ex": 4,
    "ey": 5,
    "ez": 6,
    "gxy": 7,
    "gyz": 8,
    "gxz": 9,
    "prxy": 10,
    "pryz": 11,
    "prxz": 12,
    "density": 13,
    "drydensity": 14,
    "uts": 15,
    "ucs": 16,
    "reference": 17,
}

MPa_to_Pa = 1_000_000.0


# ---------------------------------------------------------------------------
#  Helpers
# ---------------------------------------------------------------------------

def _num(val):
    """Convert a cell value to float; return NaN for None / empty."""
    if val is None:
        return np.nan
    try:
        return float(val)
    except (ValueError, TypeError):
        return np.nan


def _str(val):
    """Convert a cell value to string; return '' for None."""
    if val is None:
        return ""
    return str(val).strip()


def _read_numlist(val):
    """Parse a cell that may contain a single number or a comma-separated list
    inside square brackets, e.g. ``[0.1, 0.5, 1.0]``.
    """
    if val is None:
        return np.array([])
    # Already numeric
    try:
        return np.array([float(val)])
    except (ValueError, TypeError):
        pass
    s = str(val).strip()
    if s.startswith("["):
        s = s[1:]
    if s.endswith("]"):
        s = s[:-1]
    parts = s.split(",")
    return np.array([float(p) for p in parts])


def _read_strlist(val):
    """Parse a cell that may contain comma-separated strings."""
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    return [p.strip() for p in s.split(",")]


def _col_values(ws, col_idx, start_row, end_row):
    """Read a column slice from a worksheet and return as a list of raw values."""
    return [ws.cell(row=r, column=col_idx + 1).value for r in range(start_row, end_row + 1)]


# ---------------------------------------------------------------------------
#  Main entry point
# ---------------------------------------------------------------------------

def excel_to_blade(blade, filename: str, airfoil_dir: str = None):
    """Populate a :class:`Blade` object from a NuMAD Excel file.

    Supports both the new pyNuMAD 3-sheet format (Geometry / Components /
    Materials) and the legacy station-based format (Geometry / Materials /
    ``bend & sweep`` / SW / BOM).  The format is detected automatically by
    checking for the presence of a ``Components`` sheet.

    Parameters
    ----------
    blade : Blade
        Blade instance to populate (modified in-place).
    filename : str
        Path to the ``.xlsx`` file.
    airfoil_dir : str, optional
        Directory containing airfoil coordinate ``.txt`` files.  If *None*,
        the reader looks for a sibling ``airfoils/`` directory next to
        *filename*.

    Returns
    -------
    Blade
        The populated blade object.
    """
    wb = load_workbook(filename, read_only=True, data_only=True)

    definition = Definition()
    blade.definition = definition

    # --- resolve airfoil directory ----------------------------------------
    if airfoil_dir is None:
        candidate = os.path.join(os.path.dirname(os.path.abspath(filename)), "airfoils")
        if os.path.isdir(candidate):
            airfoil_dir = candidate

    # --- detect format and read -------------------------------------------
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    if "components" in sheet_names_lower:
        # New pyNuMAD format (3 sheets)
        _read_geometry(definition, wb[sheet_names_lower["geometry"]], airfoil_dir)
        blade.ispan = definition.ispan
        _read_materials(definition, wb[sheet_names_lower["materials"]])
        _read_components(definition, wb[sheet_names_lower["components"]])
    else:
        # Legacy station-based format
        _read_legacy_format(definition, wb, sheet_names_lower, airfoil_dir)
        blade.ispan = definition.ispan

    wb.close()
    return blade


# ---------------------------------------------------------------------------
#  Geometry sheet
# ---------------------------------------------------------------------------

def _read_geometry(definition, ws, airfoil_dir):
    """Parse the *Geometry* sheet."""

    # --- configuration flags (rows 2-4, column B) -------------------------
    flag_val = _str(ws.cell(row=2, column=2).value)
    definition.natural_offset = 1 if flag_val.upper() == "T" else 0

    flag_val = _str(ws.cell(row=3, column=2).value)
    definition.rotorspin = 1 if flag_val.upper() == "CW" else -1

    flag_val = _str(ws.cell(row=4, column=2).value)
    definition.swtwisted = 1 if flag_val.upper() == "T" else 0

    # --- determine last data row ------------------------------------------
    dr1 = _GEOM["datarow1"]  # first data row (1-indexed)
    max_row = ws.max_row
    # Read the span column to find extent of data
    span_raw = []
    for r in range(dr1, max_row + 1):
        v = _num(ws.cell(row=r, column=_GEOM["span"] + 1).value)
        if np.isnan(v):
            break
        span_raw.append(v)
    n_span = len(span_raw)
    if n_span == 0:
        raise ValueError("No span data found in Geometry sheet")
    last_row = dr1 + n_span - 1

    # --- main arrays ------------------------------------------------------
    definition.span = np.array(span_raw, dtype=float)

    definition.degreestwist = np.array(
        [_num(ws.cell(row=r, column=_GEOM["twist"] + 1).value) for r in range(dr1, last_row + 1)],
        dtype=float,
    )
    definition.chord = np.array(
        [_num(ws.cell(row=r, column=_GEOM["chord"] + 1).value) for r in range(dr1, last_row + 1)],
        dtype=float,
    )
    definition.percentthick = np.array(
        [_num(ws.cell(row=r, column=_GEOM["thick"] + 1).value) for r in range(dr1, last_row + 1)],
        dtype=float,
    )
    definition.chordoffset = np.array(
        [_num(ws.cell(row=r, column=_GEOM["offset"] + 1).value)
         for r in range(dr1, last_row + 1)],
        dtype=float,
    )
    definition.aerocenter = np.array(
        [_num(ws.cell(row=r, column=_GEOM["aerocenter"] + 1).value)
         for r in range(dr1, last_row + 1)],
        dtype=float,
    )
    definition.sweep = np.zeros(definition.span.shape)
    definition.prebend = np.zeros(definition.span.shape)

    # --- interpolate where NaN -----------------------------------------------
    for prop in ("degreestwist", "chord", "percentthick", "chordoffset", "aerocenter"):
        arr = getattr(definition, prop)
        ind = np.isnan(arr)
        if not np.any(ind):
            continue
        arr = arr.copy()
        if prop == "percentthick":
            absthick = np.multiply(arr, definition.chord) / 100.0
            i_absthick = interpolator_wrap(
                np.delete(definition.span, ind),
                np.delete(absthick, ind),
                definition.span[ind],
                "pchip",
            )
            arr[ind] = i_absthick / definition.chord[ind] * 100.0
        else:
            arr[ind] = interpolator_wrap(
                np.delete(definition.span, ind),
                np.delete(arr, ind),
                definition.span[ind],
                "pchip",
            )
        setattr(definition, prop, arr)

    # --- airfoil stations --------------------------------------------------
    # Read airfoil span locations and names (separate column range)
    afspan_raw = []
    afname_raw = []
    for r in range(dr1, max_row + 1):
        v = _num(ws.cell(row=r, column=_GEOM["afspan"] + 1).value)
        if np.isnan(v):
            break
        afspan_raw.append(v)
        afname_raw.append(_str(ws.cell(row=r, column=_GEOM["afname"] + 1).value))

    definition.stations = []
    for k, (af_span, af_name) in enumerate(zip(afspan_raw, afname_raw)):
        if af_span < np.amin(definition.span) or af_span > np.amax(definition.span):
            raise ValueError(
                f"Airfoil #{k} location ({af_span}) is outside the span range "
                f"[{definition.span[0]}, {definition.span[-1]}]"
            )
        af = _resolve_airfoil(af_name, airfoil_dir)
        if getattr(af, "_from_file", False):
            af.resample(spacing="half-cosine")
        definition.add_station(af, af_span)

    # --- interpolated span stations ----------------------------------------
    ispan_raw = []
    for r in range(dr1, max_row + 1):
        v = _num(ws.cell(row=r, column=_GEOM["ispan"] + 1).value)
        if np.isnan(v):
            break
        ispan_raw.append(v)
    definition.ispan = np.array(ispan_raw, dtype=float) if ispan_raw else definition.span.copy()


# ---------------------------------------------------------------------------
#  Materials sheet
# ---------------------------------------------------------------------------

def _read_materials(definition, ws):
    """Parse the *Materials* sheet into ``definition.materials``."""
    dr1 = _MTRL["datarow1"]
    materials_dict = {}

    for r in range(dr1, ws.max_row + 1):
        name = _str(ws.cell(row=r, column=_MTRL["name"] + 1).value)
        if not name:
            break  # end of data

        mat = Material()
        mat.name = name
        mat.type = _str(ws.cell(row=r, column=_MTRL["type"] + 1).value).lower()
        mat.layerthickness = _num(ws.cell(row=r, column=_MTRL["thickness"] + 1).value)

        mat.ex = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["ex"] + 1).value)
        mat.prxy = _num(ws.cell(row=r, column=_MTRL["prxy"] + 1).value)
        mat.density = _num(ws.cell(row=r, column=_MTRL["density"] + 1).value)
        mat.drydensity = _num(ws.cell(row=r, column=_MTRL["drydensity"] + 1).value)
        mat.uts = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["uts"] + 1).value)
        mat.ucs = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["ucs"] + 1).value)
        mat.reference = _str(ws.cell(row=r, column=_MTRL["reference"] + 1).value)

        if mat.type == "orthotropic":
            mat.ey = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["ey"] + 1).value)
            mat.ez = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["ez"] + 1).value)
            mat.gxy = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["gxy"] + 1).value)
            mat.gyz = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["gyz"] + 1).value)
            mat.gxz = MPa_to_Pa * _num(ws.cell(row=r, column=_MTRL["gxz"] + 1).value)
            mat.pryz = _num(ws.cell(row=r, column=_MTRL["pryz"] + 1).value)
            mat.prxz = _num(ws.cell(row=r, column=_MTRL["prxz"] + 1).value)
        else:
            # Isotropic: mirror properties
            mat.ey = mat.ex
            mat.ez = mat.ex
            mat.gxy = mat.ex / (2.0 * (1.0 + mat.prxy)) if mat.prxy else mat.ex
            mat.gyz = mat.gxy
            mat.gxz = mat.gxy
            mat.pryz = mat.prxy
            mat.prxz = mat.prxy

        materials_dict[mat.name] = mat

    definition.materials = materials_dict


# ---------------------------------------------------------------------------
#  Components sheet
# ---------------------------------------------------------------------------

def _read_components(definition, ws):
    """Parse the *Components* sheet into ``definition.components``."""
    # --- global sizing parameters (rows 2-5, column C) ---------------------
    definition.sparcapwidth = _num(ws.cell(row=_CMPT["paramrow1"] + 1,
                                           column=_CMPT["paramcol"] + 1).value)
    definition.leband = _num(ws.cell(row=_CMPT["paramrow1"] + 2,
                                     column=_CMPT["paramcol"] + 1).value)
    definition.teband = _num(ws.cell(row=_CMPT["paramrow1"] + 3,
                                     column=_CMPT["paramcol"] + 1).value)
    sparcapoffset = _num(ws.cell(row=_CMPT["paramrow1"] + 4,
                                 column=_CMPT["paramcol"] + 1).value)
    definition.sparcapoffset = 0.0 if np.isnan(sparcapoffset) else sparcapoffset

    # --- per-component rows ------------------------------------------------
    dr1 = _CMPT["datarow1"]
    component_dict = {}

    for r in range(dr1, ws.max_row + 1):
        name_val = _str(ws.cell(row=r, column=_CMPT["name"] + 1).value)
        if not name_val:
            break  # end of data

        comp = Component()
        comp.group = int(_num(ws.cell(row=r, column=_CMPT["group"] + 1).value))
        comp.name = name_val
        comp.materialid = _str(ws.cell(row=r, column=_CMPT["matid"] + 1).value)

        # Material id may be stored as a 1-based number referencing material
        # list position.  If numeric, resolve from definition.materials.
        try:
            mat_idx = int(float(comp.materialid))
            mat_names = list(definition.materials.keys())
            if 1 <= mat_idx <= len(mat_names):
                comp.materialid = mat_names[mat_idx - 1]
        except (ValueError, TypeError):
            pass  # already a name string

        comp.fabricangle = _read_numlist(
            ws.cell(row=r, column=_CMPT["angle"] + 1).value
        )
        comp.hpextents = _read_strlist(
            ws.cell(row=r, column=_CMPT["hpext"] + 1).value
        )
        comp.lpextents = _read_strlist(
            ws.cell(row=r, column=_CMPT["lpext"] + 1).value
        )

        cpspan = _read_numlist(ws.cell(row=r, column=_CMPT["cpspan"] + 1).value)
        cpnlay = _read_numlist(ws.cell(row=r, column=_CMPT["cpnlay"] + 1).value)
        comp.control_points = np.stack((cpspan, cpnlay), axis=1)

        imethod = _str(ws.cell(row=r, column=_CMPT["imethod"] + 1).value)
        comp.imethod = imethod if imethod else "linear"
        comp.pinnedends = 0

        if not np.any(len(comp.hpextents) == np.array([0, 1, 2])):
            raise ValueError(
                f"Component '{comp.name}': length of hpextents must be 0, 1, or 2"
            )
        if not np.any(len(comp.lpextents) == np.array([0, 1, 2])):
            raise ValueError(
                f"Component '{comp.name}': length of lpextents must be 0, 1, or 2"
            )

        component_dict[comp.name] = comp

    definition.components = component_dict


# ---------------------------------------------------------------------------
#  Airfoil resolution
# ---------------------------------------------------------------------------

def _resolve_airfoil(name: str, airfoil_dir: str | None) -> Airfoil:
    """Try to load airfoil coordinates from a file, falling back to a dummy
    circle for thick sections.

    The lookup order is:

    1. ``{airfoil_dir}/{name}.txt``
    2. ``{airfoil_dir}/{name}.dat``
    3. If not found or *airfoil_dir* is None, create a placeholder Airfoil
       with reference name only (coordinates will be ``None`` until resampled
       or assigned).
    """
    if airfoil_dir:
        for ext in (".txt", ".dat"):
            path = os.path.join(airfoil_dir, name + ext)
            if os.path.isfile(path):
                coords = _load_airfoil_txt(path)
                af = Airfoil(coords=coords, reference=name)
                af._from_file = True
                return af

    logger.warning("Airfoil '%s' not found on disk; creating placeholder.", name)
    # Return the default circular airfoil — already has valid coordinates.
    af = Airfoil(reference=name)
    af._from_file = False
    return af


def _load_airfoil_txt(path: str) -> np.ndarray:
    """Load a simple two-column (x, y) airfoil coordinate file.

    Skips header lines that cannot be parsed as floats.
    """
    rows = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("%"):
                continue
            parts = line.split()
            try:
                x, y = float(parts[0]), float(parts[1])
                rows.append((x, y))
            except (ValueError, IndexError):
                continue
    if not rows:
        raise ValueError(f"No coordinate data found in {path}")
    coords = np.array(rows, dtype=float)

    # Ensure clockwise winding starting from TE (1, 0)
    with np.errstate(divide="ignore", invalid="ignore"):
        if np.nanmean(np.gradient(np.arctan(coords[:, 1] / coords[:, 0]))) > 0:
            coords = np.flipud(coords)

    return coords


# ===========================================================================
#  Legacy NuMAD format (station-based)
# ===========================================================================

# Column layout for the legacy Geometry sheet.
_LEGACY_GEOM = {
    "datarow1": 4,     # first data row (1-indexed in Excel)
    "sta_num": 0,
    "span": 1,
    "airfoil": 2,
    "tetype": 3,
    "twist": 4,
    "chord": 5,
    "xoffset": 6,
    "aerocenter": 7,
    "notes": 8,
    "stacks_start": 9,  # first stack column (0-based)
}

# Column layout for the legacy Materials sheet (no dry-density column).
_LEGACY_MTRL = {
    "datarow1": 4,
    "id": 0,
    "name": 1,
    "type": 2,
    "thickness": 3,
    "ex": 4,
    "ey": 5,
    "ez": 6,
    "gxy": 7,
    "gyz": 8,
    "gxz": 9,
    "prxy": 10,
    "pryz": 11,
    "prxz": 12,
    "density": 13,
    "uts": 14,
    "ucs": 15,
    "reference": 16,
}

# Mapping from old NuMAD segment classification to femshell keypoint-index
# pairs on each side of the airfoil.  Keypoint ordering on the HP side:
# te(0)–e(1)–d(2)–c(3)–b(4)–a(5)–le(6).  LP is a mirror of HP.
_REGION_HP_INDICES = {
    "TE_FLAT":  (0, 1),  "TE_REINF": (1, 2),  "TE_PANEL": (2, 3),
    "SPAR":     (3, 4),  "LE_PANEL": (4, 5),  "LE":       (5, 6),
}
_REGION_LP_INDICES = {
    "LE":       (0, 1),  "LE_PANEL": (1, 2),  "SPAR":     (2, 3),
    "TE_PANEL": (3, 4),  "TE_REINF": (4, 5),  "TE_FLAT":  (5, 6),
}
_HP_INDEX_TO_LABEL = {0: "te", 1: "e", 2: "d", 3: "c", 4: "b", 5: "a", 6: "le"}
_LP_INDEX_TO_LABEL = {0: "le", 1: "a", 2: "b", 3: "c", 4: "d", 5: "e", 6: "te"}


def _classify_segment(name: str) -> str:
    """Map an old NuMAD segment name to a femshell region type."""
    tag = name.upper().replace("HP_", "").replace("LP_", "")
    if tag in ("WEBAFT", "WEBFWD", "CENTERMAINSC", "MAINSC"):
        return "SPAR"
    if tag == "LEFILLER":
        return "LE_PANEL"
    if tag in ("LEREINF", "LEPANEL"):
        return "LE"
    if tag == "TEFILLER":
        return "TE_PANEL"
    if tag == "TEREINF":
        return "TE_REINF"
    if tag == "FLAT":
        return "TE_FLAT"
    # Default: treat as spar region
    return "SPAR"


def _extent_from_regions(regions, side: str):
    """Compute keypoint extent labels for a contiguous set of region types.

    Returns a two-element list ``['label1', 'label2']`` or an empty list.
    """
    idx_map = _REGION_HP_INDICES if side == "hp" else _REGION_LP_INDICES
    label_map = _HP_INDEX_TO_LABEL if side == "hp" else _LP_INDEX_TO_LABEL

    all_indices = set()
    for r in regions:
        if r in idx_map:
            all_indices.update(idx_map[r])
    if not all_indices:
        return []
    return [label_map[min(all_indices)], label_map[max(all_indices)]]


# ---------------------------------------------------------------------------
#  Legacy format orchestration
# ---------------------------------------------------------------------------

def _read_legacy_format(definition, wb, sheet_names_lower, airfoil_dir):
    """Read old NuMAD station-based format into *definition*."""

    ws_geom = wb[sheet_names_lower["geometry"]]
    ws_mtrl = wb[sheet_names_lower["materials"]]

    # --- materials --------------------------------------------------------
    _read_legacy_materials(definition, ws_mtrl)

    # --- geometry + station data ------------------------------------------
    legacy = _read_legacy_geometry(definition, ws_geom, airfoil_dir)

    # The legacy format stores chordoffset as the pitch-axis fraction
    # measured from the LE — the same convention used by the YAML reader.
    # ``natural_offset=0`` avoids the extra xoffset shift that would move
    # the twist-rotation centre away from the pitch axis.
    definition.natural_offset = 0

    # --- bend & sweep -----------------------------------------------------
    if "bend & sweep" in sheet_names_lower:
        _read_legacy_bend_sweep(definition, wb[sheet_names_lower["bend & sweep"]])

    # --- spar-cap / shear-web sizing from SW sheet ------------------------
    ws_sw = wb[sheet_names_lower["sw"]] if "sw" in sheet_names_lower else None
    _build_legacy_sizing(definition, legacy, ws_sw)

    # --- convert stacks + segments to components --------------------------
    _build_legacy_components(definition, legacy)


# ---------------------------------------------------------------------------
#  Legacy Materials sheet
# ---------------------------------------------------------------------------

def _read_legacy_materials(definition, ws):
    """Parse the legacy *Materials* sheet (data starts at row 4, no
    dry-density column)."""
    L = _LEGACY_MTRL
    dr1 = L["datarow1"]
    materials_dict = {}

    for r in range(dr1, ws.max_row + 1):
        name = _str(ws.cell(row=r, column=L["name"] + 1).value)
        if not name:
            break

        mat = Material()
        mat.name = name
        mat.type = _str(ws.cell(row=r, column=L["type"] + 1).value).lower()
        mat.layerthickness = _num(ws.cell(row=r, column=L["thickness"] + 1).value)

        mat.ex = MPa_to_Pa * _num(ws.cell(row=r, column=L["ex"] + 1).value)
        mat.prxy = _num(ws.cell(row=r, column=L["prxy"] + 1).value)
        mat.density = _num(ws.cell(row=r, column=L["density"] + 1).value)
        mat.drydensity = np.nan  # not present in legacy format
        mat.uts = MPa_to_Pa * _num(ws.cell(row=r, column=L["uts"] + 1).value)
        mat.ucs = MPa_to_Pa * _num(ws.cell(row=r, column=L["ucs"] + 1).value)
        mat.reference = _str(ws.cell(row=r, column=L["reference"] + 1).value)

        if mat.type == "orthotropic":
            mat.ey = MPa_to_Pa * _num(ws.cell(row=r, column=L["ey"] + 1).value)
            mat.ez = MPa_to_Pa * _num(ws.cell(row=r, column=L["ez"] + 1).value)
            mat.gxy = MPa_to_Pa * _num(ws.cell(row=r, column=L["gxy"] + 1).value)
            mat.gyz = MPa_to_Pa * _num(ws.cell(row=r, column=L["gyz"] + 1).value)
            mat.gxz = MPa_to_Pa * _num(ws.cell(row=r, column=L["gxz"] + 1).value)
            mat.pryz = _num(ws.cell(row=r, column=L["pryz"] + 1).value)
            mat.prxz = _num(ws.cell(row=r, column=L["prxz"] + 1).value)
        else:
            mat.ey = mat.ex
            mat.ez = mat.ex
            mat.gxy = mat.ex / (2.0 * (1.0 + mat.prxy)) if mat.prxy else mat.ex
            mat.gyz = mat.gxy
            mat.gxz = mat.gxy
            mat.pryz = mat.prxy
            mat.prxz = mat.prxy

        materials_dict[mat.name] = mat

    definition.materials = materials_dict


# ---------------------------------------------------------------------------
#  Legacy Geometry sheet
# ---------------------------------------------------------------------------

def _read_legacy_geometry(definition, ws, airfoil_dir):
    """Parse the legacy *Geometry* sheet.

    Returns a dict with the station-based layout data needed by the
    component builder: stack names, material IDs, layer counts, segment
    names, DP positions, included-stacks per segment, and shear-web info.

    Uses bulk row reading via ``iter_rows`` to avoid slow random cell
    access on wide sheets with formatting artefacts.
    """
    G = _LEGACY_GEOM
    dr1 = G["datarow1"]  # first data row (1-indexed)

    # --- pre-read all needed rows (limit to 80 columns) -------------------
    MAX_COL = 80
    all_rows = {}  # row_number → tuple of values
    for idx, row_vals in enumerate(
            ws.iter_rows(min_row=1, max_col=MAX_COL, values_only=True)):
        r = idx + 1  # 1-based row number
        vals = tuple(row_vals)
        if len(vals) < MAX_COL:
            vals += (None,) * (MAX_COL - len(vals))
        all_rows[r] = vals

    def _cv(row, col0):
        """Get cell value at 0-based column *col0* from pre-read row."""
        vals = all_rows.get(row)
        if vals is None or col0 >= len(vals):
            return None
        return vals[col0]

    # --- header rows (1-3) -----------------------------------------------
    n_shear_webs = int(_num(_cv(1, 5)))      # row 1, col F (0-based 5)
    n_stacks = int(_num(_cv(1, 7)))           # row 1, col H (0-based 7)

    n_segments = int(_num(_cv(2, 7)))         # row 2, col H
    stack_mat_ids = [int(_num(_cv(2, G["stacks_start"] + i)))
                     for i in range(n_stacks)]

    stack_names = [_str(_cv(3, G["stacks_start"] + i))
                   for i in range(n_stacks)]
    seg_col0 = G["stacks_start"] + n_stacks + 1
    segment_names = [_str(_cv(3, seg_col0 + i)) for i in range(n_segments)]

    # --- determine data extent (count stations) ---------------------------
    span_raw = []
    for r in sorted(all_rows):
        if r < dr1:
            continue
        v = _num(_cv(r, G["span"]))
        if np.isnan(v):
            break
        span_raw.append(v)
    n_sta = len(span_raw)
    if n_sta == 0:
        raise ValueError("No span data found in legacy Geometry sheet")
    last_row = dr1 + n_sta - 1

    # --- geometry arrays --------------------------------------------------
    definition.span = np.array(span_raw, dtype=float)
    definition.degreestwist = np.array(
        [_num(_cv(r, G["twist"])) for r in range(dr1, last_row + 1)], dtype=float)
    definition.chord = np.array(
        [_num(_cv(r, G["chord"])) for r in range(dr1, last_row + 1)], dtype=float)
    definition.chordoffset = np.array(
        [_num(_cv(r, G["xoffset"])) for r in range(dr1, last_row + 1)], dtype=float)
    definition.aerocenter = np.array(
        [_num(_cv(r, G["aerocenter"])) for r in range(dr1, last_row + 1)], dtype=float)

    definition.percentthick = np.full(n_sta, np.nan)
    definition.sweep = np.zeros(n_sta)
    definition.prebend = np.zeros(n_sta)

    te_types = [_str(_cv(r, G["tetype"])).lower()
                for r in range(dr1, last_row + 1)]
    definition.te_type = te_types

    # --- interpolate NaN geometry -----------------------------------------
    for prop in ("degreestwist", "chord", "chordoffset", "aerocenter"):
        arr = getattr(definition, prop)
        ind = np.isnan(arr)
        if not np.any(ind):
            continue
        arr = arr.copy()
        arr[ind] = interpolator_wrap(
            np.delete(definition.span, ind),
            np.delete(arr, ind),
            definition.span[ind],
            "pchip",
        )
        setattr(definition, prop, arr)

    # --- airfoil stations -------------------------------------------------
    definition.stations = []
    for r in range(dr1, last_row + 1):
        af_name = _str(_cv(r, G["airfoil"]))
        af_span = definition.span[r - dr1]
        af = _resolve_airfoil(af_name, airfoil_dir)
        if getattr(af, "_from_file", False):
            af.resample(spacing="half-cosine")
            coords = af.coordinates
            if coords is not None and len(coords) > 2:
                max_thick = np.max(coords[:, 1]) - np.min(coords[:, 1])
                definition.percentthick[r - dr1] = max_thick * 100.0
        definition.add_station(af, af_span)

    definition.ispan = definition.span.copy()

    # Fill missing percentthick
    pt = definition.percentthick
    if np.any(np.isnan(pt)):
        known = ~np.isnan(pt)
        if np.any(known):
            pt[~known] = interpolator_wrap(
                definition.span[known], pt[known],
                definition.span[~known], "pchip")
        else:
            pt[:] = np.linspace(100, 21, n_sta)
        definition.percentthick = pt

    # --- station-based layup data -----------------------------------------
    layer_counts = np.zeros((n_sta, n_stacks))
    for k in range(n_sta):
        r = dr1 + k
        for j in range(n_stacks):
            layer_counts[k, j] = _num(_cv(r, G["stacks_start"] + j))
    layer_counts = np.nan_to_num(layer_counts, nan=0.0)

    dp_col0 = G["stacks_start"] + n_stacks  # 0-based column of first DP
    dp_positions = np.zeros((n_sta, n_segments + 1))
    for k in range(n_sta):
        r = dr1 + k
        for j in range(n_segments + 1):
            dp_positions[k, j] = _num(_cv(r, dp_col0 + j))

    # incl_stacks -- use first data row as representative
    incl_col0 = G["stacks_start"] + n_stacks + (n_segments + 1) + n_segments + 1
    incl_stacks = []
    for j in range(n_segments):
        raw = _str(_cv(dr1, incl_col0 + j))
        if raw:
            incl_stacks.append([int(x) for x in raw.split(",")])
        else:
            incl_stacks.append([])

    # Shear web stacks and DP connections
    sw_col0 = incl_col0 + n_segments
    sw_stacks_lists = []
    sw_dp_conn = []
    for w in range(n_shear_webs):
        stk_raw = None
        dp_raw = None
        for k in range(n_sta):
            r = dr1 + k
            s = _str(_cv(r, sw_col0 + 2 * w))
            d = _str(_cv(r, sw_col0 + 1 + 2 * w))
            if s:
                stk_raw = s
                dp_raw = d
                break
        if stk_raw:
            sw_stacks_lists.append([int(x) for x in stk_raw.split(",")])
        else:
            sw_stacks_lists.append([])
        if dp_raw:
            sw_dp_conn.append([int(x) for x in dp_raw.split(",")])
        else:
            sw_dp_conn.append([])

    return {
        "n_stacks": n_stacks,
        "n_segments": n_segments,
        "n_shear_webs": n_shear_webs,
        "n_sta": n_sta,
        "stack_names": stack_names,
        "stack_mat_ids": stack_mat_ids,
        "segment_names": segment_names,
        "layer_counts": layer_counts,
        "dp_positions": dp_positions,
        "incl_stacks": incl_stacks,
        "sw_stacks_lists": sw_stacks_lists,
        "sw_dp_conn": sw_dp_conn,
    }

# ---------------------------------------------------------------------------
#  Legacy bend & sweep sheet
# ---------------------------------------------------------------------------

def _read_legacy_bend_sweep(definition, ws):
    """Parse the legacy *bend & sweep* sheet for prebend and sweep data."""
    # Layout: row 1: headers (Pre-Bend, method, type / Sweep, method, type)
    #          row 2: type info
    #          row 3: column sub-headers
    #          row 4+: data columns: span, disp, slope (prebend) | span, disp, slope (sweep)
    dr1 = 4  # first data row (1-indexed)

    prebend_vals = []
    prebend_spans = []
    sweep_vals = []
    sweep_spans = []

    for r in range(dr1, ws.max_row + 1):
        pb_span = _num(ws.cell(row=r, column=1).value)
        if np.isnan(pb_span):
            break
        prebend_spans.append(pb_span)
        prebend_vals.append(_num(ws.cell(row=r, column=2).value))

    for r in range(dr1, ws.max_row + 1):
        sw_span = _num(ws.cell(row=r, column=4).value)
        if np.isnan(sw_span):
            break
        sweep_spans.append(sw_span)
        sweep_vals.append(_num(ws.cell(row=r, column=5).value))

    # Interpolate onto definition span stations
    if prebend_spans:
        pb_s = np.array(prebend_spans)
        pb_v = np.nan_to_num(np.array(prebend_vals), nan=0.0)
        definition.prebend = interpolator_wrap(
            pb_s, pb_v, definition.span, "pchip")
    if sweep_spans:
        sw_s = np.array(sweep_spans)
        sw_v = np.nan_to_num(np.array(sweep_vals), nan=0.0)
        definition.sweep = interpolator_wrap(
            sw_s, sw_v, definition.span, "pchip")


# ---------------------------------------------------------------------------
#  Legacy sizing parameters (sparcapwidth, leband, teband)
# ---------------------------------------------------------------------------

def _build_legacy_sizing(definition, legacy, ws_sw):
    """Compute sizing parameters from legacy data and the SW sheet.

    Sets ``definition.sparcapwidth_hp``, ``definition.sparcapwidth_lp``,
    ``definition.sparcapoffset_hp``, ``definition.sparcapoffset_lp``,
    ``definition.leband``, ``definition.teband`` as arrays over
    ``definition.span``.
    """
    n_sta = legacy["n_sta"]
    dp = legacy["dp_positions"]        # [n_sta, n_segments+1]
    chord = definition.chord            # [n_sta]
    n_seg = legacy["n_segments"]
    seg_names = legacy["segment_names"]

    # ---- spar-cap width (use SW sheet if available) ----------------------
    if ws_sw is not None:
        # Column 9 (1-indexed) = "Sparcap linear width" in meters
        scw = []
        for r in range(4, 4 + n_sta):
            v = _num(ws_sw.cell(row=r, column=9).value)
            scw.append(v if not np.isnan(v) else 0.0)
        scw = np.array(scw) * 1000.0  # m → mm
    else:
        # Approximate from DPs: spar region spans from c to b
        # Find spar-boundary DPs heuristically
        hp_c_idx = None  # DP index of the TE-side spar boundary on HP
        hp_b_idx = None  # DP index of the LE-side spar boundary on HP
        for j, nm in enumerate(seg_names):
            cls = _classify_segment(nm)
            if nm.upper().startswith("HP"):
                if cls == "SPAR" and hp_c_idx is None:
                    hp_c_idx = j  # first SPAR segment starts at DP j+1? no — segment j is between DP j and DP j+1
                if cls == "SPAR":
                    hp_b_idx = j + 1  # last SPAR segment ends at DP j+2
        # +1 offset: segment j lies between DP[j] and DP[j+1]
        if hp_c_idx is not None and hp_b_idx is not None:
            scw = np.abs(dp[:, hp_b_idx + 1] - dp[:, hp_c_idx + 1]) * chord * 1000.0
        else:
            scw = np.full(n_sta, 600.0)  # fallback 600 mm

    definition.sparcapwidth_hp = scw.copy()
    definition.sparcapwidth_lp = scw.copy()
    definition.sparcapoffset_hp = np.zeros(n_sta)
    definition.sparcapoffset_lp = np.zeros(n_sta)

    # ---- leband and teband -----------------------------------------------
    # Find the DP indices corresponding to keypoints 'a' and 'd' on HP side
    # 'a' ≈ boundary between the last HP non-LE segment and LE
    # 'd' ≈ boundary between TE-REINF and TE-PANEL on HP
    hp_a_dp_idx = None
    hp_d_dp_idx = None
    for j, nm in enumerate(seg_names):
        if not nm.upper().startswith("HP"):
            continue
        cls = _classify_segment(nm)
        if cls == "LE_PANEL" and hp_a_dp_idx is None:
            # LE_PANEL = segment between b and a.
            # This segment, say index j, sits between DP[j] and DP[j+1].
            # Keypoint 'a' = DP[j+1] (the LE-side end of LE_PANEL).
            hp_a_dp_idx = j + 1
        if cls == "TE_PANEL" and hp_d_dp_idx is None:
            # TE_PANEL = segment between d and c.
            # Keypoint 'd' = DP[j] (the TE-side end of TE_PANEL).
            hp_d_dp_idx = j

    if hp_a_dp_idx is not None:
        # leband = distance from LE to keypoint 'a' ≈ |DP_a| * chord * 1000
        definition.leband = np.abs(dp[:, hp_a_dp_idx]) * chord * 1000.0
    else:
        definition.leband = np.full(n_sta, 50.0)

    if hp_d_dp_idx is not None:
        # teband = distance from TE to keypoint 'd' ≈ (1 - |DP_d|) * chord * 1000
        definition.teband = (1.0 - np.abs(dp[:, hp_d_dp_idx])) * chord * 1000.0
    else:
        definition.teband = np.full(n_sta, 200.0)

    # Interpolate NaN values (legacy format only defines DPs at key stations)
    span = definition.span
    for attr in ("leband", "teband"):
        arr = getattr(definition, attr)
        nans = np.isnan(arr)
        if np.any(nans) and np.any(~nans):
            arr[nans] = interpolator_wrap(
                span[~nans], arr[~nans], span[nans], "pchip")
            setattr(definition, attr, arr)

    # Clamp to avoid zero-width bands
    definition.leband = np.maximum(definition.leband, 1.0)
    definition.teband = np.maximum(definition.teband, 1.0)


# ---------------------------------------------------------------------------
#  Legacy component builder
# ---------------------------------------------------------------------------

def _build_legacy_components(definition, legacy):
    """Convert station-based stacks/segments to Component objects.

    Each unique (stack-position, stack-id) pair across the segment stacking
    sequences becomes one blade-surface component.  Shear webs become
    additional components with ``group > 0``.
    """
    n_sta = legacy["n_sta"]
    n_stacks = legacy["n_stacks"]
    stack_names = legacy["stack_names"]
    stack_mat_ids = legacy["stack_mat_ids"]
    segment_names = legacy["segment_names"]
    incl_stacks = legacy["incl_stacks"]
    layer_counts = legacy["layer_counts"]
    span = definition.span
    ndspan = span / span[-1]  # normalised [0, 1] for control points

    mat_names = list(definition.materials.keys())

    # --- analyse stacking sequences to determine component extents --------
    # For each (position-in-sequence, stack_id) find which region types it
    # covers on HP and LP sides.
    entry_regions = {}  # key=(position, stack_id) → {'hp': set, 'lp': set}
    for seg_idx, seg_name in enumerate(segment_names):
        side = "hp" if seg_name.upper().startswith("HP") else "lp"
        region = _classify_segment(seg_name)
        seq = incl_stacks[seg_idx]
        for pos, stk_id in enumerate(seq):
            key = (pos, stk_id)
            if key not in entry_regions:
                entry_regions[key] = {"hp": set(), "lp": set()}
            entry_regions[key][side].add(region)

    # Determine a canonical ordering: group entries by their first position
    # in the stacking sequence so that outer layers come first.
    max_seq_len = max(len(s) for s in incl_stacks) if incl_stacks else 0
    ordered_entries = sorted(entry_regions.keys(), key=lambda k: (k[0], k[1]))

    # Track how many times each stack_id has been used (for naming)
    stack_id_count = {}

    component_dict = {}
    for pos, stk_id in ordered_entries:
        regions = entry_regions[(pos, stk_id)]
        hp_ext = _extent_from_regions(regions["hp"], "hp")
        lp_ext = _extent_from_regions(regions["lp"], "lp")
        if not hp_ext and not lp_ext:
            continue

        # Material name from the material ID
        mat_id_idx = stk_id - 1  # 1-based → 0-based
        if 0 <= mat_id_idx < len(stack_mat_ids):
            mat_ref = stack_mat_ids[mat_id_idx] - 1  # material ID (1-based)
            if 0 <= mat_ref < len(mat_names):
                mat_name = mat_names[mat_ref]
            else:
                mat_name = mat_names[0]
        else:
            mat_name = mat_names[0]

        # Component name
        base_name = stack_names[mat_id_idx] if 0 <= mat_id_idx < len(stack_names) else f"Stack{stk_id}"
        stack_id_count[stk_id] = stack_id_count.get(stk_id, 0) + 1
        if stack_id_count[stk_id] > 1:
            comp_name = f"{base_name}_inner{stack_id_count[stk_id] - 1}"
        else:
            comp_name = base_name

        comp = Component()
        comp.group = 0
        comp.name = comp_name
        comp.materialid = mat_name
        comp.fabricangle = np.array([0.0])
        comp.hpextents = hp_ext
        comp.lpextents = lp_ext
        comp.imethod = "pchip"
        comp.pinnedends = 0

        # Control points: (normalised span, n_layers) at each station
        stk_col = mat_id_idx  # column in layer_counts
        n_layers = layer_counts[:, stk_col]
        # Interpolate NaN values
        n_layers = np.nan_to_num(n_layers, nan=0.0)
        n_layers = np.round(n_layers).astype(float)
        comp.control_points = np.column_stack((ndspan, n_layers))

        component_dict[comp_name] = comp

    # --- shear web components ---------------------------------------------
    sw_stacks_lists = legacy["sw_stacks_lists"]
    n_sw = legacy["n_shear_webs"]
    web_keypoints = ["b", "c"]  # SW1 at 'b', SW2 at 'c' (standard convention)

    for w in range(n_sw):
        if not sw_stacks_lists[w]:
            continue
        sw_seq = sw_stacks_lists[w]
        kp = web_keypoints[w] if w < len(web_keypoints) else "b"

        stk_count = {}
        for pos, stk_id in enumerate(sw_seq):
            mat_id_idx = stk_id - 1
            if 0 <= mat_id_idx < len(stack_mat_ids):
                mat_ref = stack_mat_ids[mat_id_idx] - 1
                mat_name = mat_names[mat_ref] if 0 <= mat_ref < len(mat_names) else mat_names[0]
            else:
                mat_name = mat_names[0]

            base_name = stack_names[mat_id_idx] if 0 <= mat_id_idx < len(stack_names) else f"Stack{stk_id}"
            stk_count[stk_id] = stk_count.get(stk_id, 0) + 1
            suffix = f"_sw{w + 1}"
            if stk_count[stk_id] > 1:
                suffix += f"_{stk_count[stk_id]}"
            comp_name = f"{base_name}{suffix}"

            comp = Component()
            comp.group = w + 1
            comp.name = comp_name
            comp.materialid = mat_name
            comp.fabricangle = np.array([0.0])
            comp.hpextents = [kp]
            comp.lpextents = [kp]
            comp.imethod = "pchip"
            comp.pinnedends = 0

            stk_col = mat_id_idx
            n_layers = np.nan_to_num(layer_counts[:, stk_col], nan=0.0)
            n_layers = np.round(n_layers).astype(float)
            comp.control_points = np.column_stack((ndspan, n_layers))

            component_dict[comp_name] = comp

    definition.components = component_dict
    logger.info("Legacy format: created %d components (%d blade + %d web)",
                len(component_dict),
                sum(1 for c in component_dict.values() if c.group == 0),
                sum(1 for c in component_dict.values() if c.group > 0))

